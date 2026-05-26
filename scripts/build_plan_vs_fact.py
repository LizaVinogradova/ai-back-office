from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


WORKBOOK = Path("monthly_workbook.xlsx")
OUTPUT_XLSX = Path("plan-vs-fact.xlsx")
OUTPUT_MD = Path("plan-vs-fact-summary.md")

PROJECTS = {
    "PRJ-2026-001": "Весенние чтения",
    "PRJ-2026-002": "Региональная программа — Сибирь",
    "PRJ-2026-003": "Цифровая платформа 2.0",
    "PRJ-2026-004": "B2B со Сбером",
    "PRJ-2026-005": "Аудиокниги — летний релиз",
}


def money(value: float) -> str:
    return f"{value:,.0f}".replace(",", " ")


def pct(value: float) -> str:
    return f"{value:.1%}"


def plan_columns(df: pd.DataFrame) -> list[str]:
    return [col for col in df.columns if isinstance(col, str) and col.startswith("План ")]


def month_from_plan_col(col: str) -> str:
    match = re.search(r"(\d{4}-\d{2})", col)
    if not match:
        raise ValueError(f"Cannot extract month from {col!r}")
    return match.group(1)


def article_from_1c(value: str) -> str:
    text = str(value)
    return text.split("—", 1)[1].strip() if "—" in text else text.strip()


def read_sources() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    plan_income = pd.read_excel(WORKBOOK, sheet_name="План - доходы")
    plan_expenses = pd.read_excel(WORKBOOK, sheet_name="План - расходы")
    actual_income = pd.read_excel(WORKBOOK, sheet_name="Факт - доходы")

    actual_expense_frames = []
    for path in sorted(Path(".").glob("expenses_fact_1c_2026-*.xlsx")):
        df = pd.read_excel(path)
        df["Дата"] = pd.to_datetime(df["Дата"], dayfirst=True)
        df["Месяц"] = df["Дата"].dt.strftime("%Y-%m")
        df["Project ID"] = df["Код проекта"]
        df["Статья"] = df["Статья (1С)"].map(article_from_1c)
        df["Факт расходов, тыс ₽"] = df["Сумма, ₽"] / 1000
        actual_expense_frames.append(df)

    actual_expenses = pd.concat(actual_expense_frames, ignore_index=True)

    actual_income["Дата"] = pd.to_datetime(actual_income["Дата"], dayfirst=True)
    actual_income["Месяц"] = actual_income["Дата"].dt.strftime("%Y-%m")
    actual_income["Факт доходов, тыс ₽"] = actual_income["Сумма, ₽"] / 1000

    months = [month_from_plan_col(col) for col in plan_columns(plan_expenses)]

    income_rows = []
    for _, row in plan_income.iterrows():
        for col in plan_columns(plan_income):
            income_rows.append(
                {
                    "Project ID": row["Project ID"],
                    "Проект": PROJECTS.get(row["Project ID"], row["Project ID"]),
                    "Месяц": month_from_plan_col(col),
                    "Источник": row["Источник"],
                    "План доходов, тыс ₽": float(row[col]),
                }
            )
    income_detail = pd.DataFrame(income_rows)
    actual_income_grouped = (
        actual_income.groupby(["Project ID", "Месяц"], as_index=False)["Факт доходов, тыс ₽"].sum()
    )
    income_detail = income_detail.merge(actual_income_grouped, on=["Project ID", "Месяц"], how="left")
    income_detail["Факт доходов, тыс ₽"] = income_detail["Факт доходов, тыс ₽"].fillna(0)

    expense_rows = []
    for _, row in plan_expenses.iterrows():
        for col in plan_columns(plan_expenses):
            expense_rows.append(
                {
                    "Project ID": row["Project ID"],
                    "Проект": PROJECTS.get(row["Project ID"], row["Project ID"]),
                    "Месяц": month_from_plan_col(col),
                    "Статья": row["Статья"],
                    "План расходов, тыс ₽": float(row[col]),
                }
            )
    expense_detail = pd.DataFrame(expense_rows)
    actual_expense_grouped = (
        actual_expenses.groupby(["Project ID", "Месяц", "Статья"], as_index=False)["Факт расходов, тыс ₽"].sum()
    )

    # Keep all planned rows, plus any 1C fact rows whose article is not in the plan.
    detail = expense_detail.merge(
        actual_expense_grouped,
        on=["Project ID", "Месяц", "Статья"],
        how="outer",
    )
    detail["Проект"] = detail["Проект"].fillna(detail["Project ID"].map(PROJECTS))
    detail["План расходов, тыс ₽"] = detail["План расходов, тыс ₽"].fillna(0)
    detail["Факт расходов, тыс ₽"] = detail["Факт расходов, тыс ₽"].fillna(0)
    detail = detail[detail["Project ID"].isin(PROJECTS)].sort_values(["Project ID", "Месяц", "Статья"])

    return income_detail, detail, pd.DataFrame({"Месяц": months})


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def autosize(ws, max_width: int = 42) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, max_width)
        ws.column_dimensions[letter].width = max(width, 12)


def build_workbook(income_detail: pd.DataFrame, expense_detail: pd.DataFrame) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Сводка"
    ws_detail = wb.create_sheet("Детализация")
    ws_income = wb.create_sheet("Данные_доходы")
    ws_income.sheet_state = "hidden"

    summary_headers = [
        "Project ID",
        "Проект",
        "План доходов, тыс ₽",
        "Факт доходов, тыс ₽",
        "План расходов, тыс ₽",
        "Факт расходов, тыс ₽",
        "EBITDA-план, тыс ₽",
        "EBITDA-факт, тыс ₽",
        "Отклонение EBITDA, тыс ₽",
        "Отклонение EBITDA, %",
    ]
    ws_summary.append(summary_headers)

    detail_headers = ["Project ID", "Проект", "Месяц", "Статья", "План расходов, тыс ₽", "Факт расходов, тыс ₽"]
    ws_detail.append(detail_headers)
    for _, row in expense_detail.iterrows():
        ws_detail.append([row[col] for col in detail_headers])

    income_headers = ["Project ID", "Проект", "Месяц", "Источник", "План доходов, тыс ₽", "Факт доходов, тыс ₽"]
    ws_income.append(income_headers)
    for _, row in income_detail.iterrows():
        ws_income.append([row[col] for col in income_headers])

    for idx, (project_id, project_name) in enumerate(PROJECTS.items(), start=2):
        ws_summary.cell(idx, 1, project_id)
        ws_summary.cell(idx, 2, project_name)
        ws_summary.cell(idx, 3, f'=SUMIFS(Данные_доходы!E:E,Данные_доходы!A:A,A{idx})')
        ws_summary.cell(idx, 4, f'=SUMIFS(Данные_доходы!F:F,Данные_доходы!A:A,A{idx})')
        ws_summary.cell(idx, 5, f'=SUMIFS(Детализация!E:E,Детализация!A:A,A{idx})')
        ws_summary.cell(idx, 6, f'=SUMIFS(Детализация!F:F,Детализация!A:A,A{idx})')
        ws_summary.cell(idx, 7, f"=C{idx}-E{idx}")
        ws_summary.cell(idx, 8, f"=D{idx}-F{idx}")
        ws_summary.cell(idx, 9, f"=H{idx}-G{idx}")
        ws_summary.cell(idx, 10, f"=IFERROR(I{idx}/ABS(G{idx}),0)")

    total_row = len(PROJECTS) + 2
    ws_summary.cell(total_row, 1, "Итого")
    ws_summary.cell(total_row, 2, "")
    for col in range(3, 10):
        letter = get_column_letter(col)
        ws_summary.cell(total_row, col, f"=SUM({letter}2:{letter}{total_row - 1})")
    ws_summary.cell(total_row, 10, f"=IFERROR(I{total_row}/ABS(G{total_row}),0)")

    style_sheet(ws_summary)
    style_sheet(ws_detail)
    style_sheet(ws_income)

    for ws in [ws_summary, ws_detail, ws_income]:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        autosize(ws)

    for row in range(2, total_row + 1):
        for col in range(3, 10):
            ws_summary.cell(row, col).number_format = '# ##0'
        ws_summary.cell(row, 10).number_format = '0.0%'
    for row in range(2, ws_detail.max_row + 1):
        ws_detail.cell(row, 5).number_format = '# ##0'
        ws_detail.cell(row, 6).number_format = '# ##0'
    for row in range(2, ws_income.max_row + 1):
        ws_income.cell(row, 5).number_format = '# ##0'
        ws_income.cell(row, 6).number_format = '# ##0'

    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    ws_summary.conditional_formatting.add(f"I2:I{total_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green))
    ws_summary.conditional_formatting.add(f"I2:I{total_row}", CellIsRule(operator="lessThan", formula=["0"], fill=red))
    ws_summary.conditional_formatting.add(f"J2:J{total_row}", CellIsRule(operator="greaterThanOrEqual", formula=["0"], fill=green))
    ws_summary.conditional_formatting.add(f"J2:J{total_row}", CellIsRule(operator="lessThan", formula=["0"], fill=red))

    ws_summary.row_dimensions[1].height = 34
    ws_detail.row_dimensions[1].height = 34
    ws_summary.sheet_view.showGridLines = False
    ws_detail.sheet_view.showGridLines = False

    wb.save(OUTPUT_XLSX)


def build_summary_md(income_detail: pd.DataFrame, expense_detail: pd.DataFrame) -> None:
    rows = []
    for project_id, project_name in PROJECTS.items():
        plan_income = float(income_detail.loc[income_detail["Project ID"] == project_id, "План доходов, тыс ₽"].sum())
        fact_income = float(income_detail.loc[income_detail["Project ID"] == project_id, "Факт доходов, тыс ₽"].sum())
        plan_expense = float(expense_detail.loc[expense_detail["Project ID"] == project_id, "План расходов, тыс ₽"].sum())
        fact_expense = float(expense_detail.loc[expense_detail["Project ID"] == project_id, "Факт расходов, тыс ₽"].sum())
        ebitda_plan = plan_income - plan_expense
        ebitda_fact = fact_income - fact_expense
        delta = ebitda_fact - ebitda_plan
        delta_pct = delta / abs(ebitda_plan) if ebitda_plan else 0
        rows.append([project_id, project_name, plan_income, fact_income, plan_expense, fact_expense, ebitda_plan, ebitda_fact, delta, delta_pct])

    totals = ["Итого", "", *[sum(row[i] for row in rows) for i in range(2, 9)]]
    totals.append(totals[-1] / abs(totals[-3]) if totals[-3] else 0)

    lines = [
        "# EBITDA план vs факт Q1 2026",
        "",
        "| Project ID | Проект | План доходов, тыс ₽ | Факт доходов, тыс ₽ | План расходов, тыс ₽ | Факт расходов, тыс ₽ | EBITDA-план, тыс ₽ | EBITDA-факт, тыс ₽ | Отклонение EBITDA, тыс ₽ | Отклонение EBITDA, % |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    row[0],
                    row[1],
                    money(row[2]),
                    money(row[3]),
                    money(row[4]),
                    money(row[5]),
                    money(row[6]),
                    money(row[7]),
                    money(row[8]),
                    pct(row[9]),
                ]
            )
            + " |"
        )
    lines.append(
        "| "
        + " | ".join(
            [
                totals[0],
                totals[1],
                money(totals[2]),
                money(totals[3]),
                money(totals[4]),
                money(totals[5]),
                money(totals[6]),
                money(totals[7]),
                money(totals[8]),
                pct(totals[9]),
            ]
        )
        + " |"
    )
    lines.append("")
    OUTPUT_MD.write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    income_detail, expense_detail, _months = read_sources()
    build_workbook(income_detail, expense_detail)
    build_summary_md(income_detail, expense_detail)


if __name__ == "__main__":
    main()
